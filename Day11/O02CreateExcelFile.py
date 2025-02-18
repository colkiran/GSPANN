"""
pip install openpyexcel

Workbook - excel file
WorkSheet - excel sheets
cells  - excel sheet - cells
"""

from openpyxl import Workbook
from datetime import datetime

wb = Workbook()

ws = wb.active

ws.title = "MyExcelSheet01"

ws['B9'] = "Hello World"
ws['D9'] = 85450.75
ws['F9'] = datetime.now()
ws['H9'].value = "=column()"

wb.save("FirstExcel.xlsx")
