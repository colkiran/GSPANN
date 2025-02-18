
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart3D
from openpyxl.chart.label import DataLabelList

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Chart")

wb.active = wb['Chart']

ws = wb.active

data = [
    ('Products', 'Sales'),
    ('Pepsi', 450),
    ('Coke', 380),
    ('Sprite', 300),
    ("Mirinda", 340),
    ("Thumbs up", 485),
    ("fanta", 285),
    ("slice", 250)
]

for row in data:
    ws.append(row)

dataRef = Reference(ws, min_row=2, min_col=2, max_row=8)
labelRef = Reference(ws, min_row=2, min_col=1, max_row=8)

chart = BarChart3D()
chart.add_data(dataRef)
chart.set_categories(labelRef)
chart.title = 'Baverage Sales'
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True

ws.add_chart(chart, "E9")
wb.save("FirstExcel.xlsx")










